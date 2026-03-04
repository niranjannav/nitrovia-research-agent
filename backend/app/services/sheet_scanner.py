"""Sheet scanner for Excel files.

Reads only the structure of an Excel workbook (sheet names, headers, sample rows,
column types) without loading all data. The resulting SheetSummary JSON is compact
(~5-10 KB) and suitable for sending to an LLM for schema inference.

This is the "understanding layer" — it describes what is in each tab so Claude can
map columns to standard analytics fields regardless of how the file is formatted.
"""

import logging
import re
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# How many data rows to include as sample
SAMPLE_ROW_COUNT = 5
# How many rows to scan for type inference
TYPE_SCAN_ROWS = 50
# Max columns to describe per sheet (avoids token explosion on wide sheets)
MAX_COLS_DESCRIBED = 60
# How many formula examples to surface per sheet
MAX_FORMULA_EXAMPLES = 3


@dataclass
class ColumnDescription:
    """Description of a single Excel column."""

    index: int           # 0-based column index
    name: str            # Header value
    detected_type: str   # 'text' | 'numeric' | 'date' | 'formula' | 'empty' | 'mixed'
    sample_values: list[Any] = field(default_factory=list)
    non_empty_count: int = 0
    formula_example: str | None = None  # First formula string found, if any


@dataclass
class SheetDescription:
    """Compact description of a single Excel sheet."""

    name: str
    row_count: int
    col_count: int
    header_row_index: int   # 1-based row index of the detected header
    has_formulas: bool
    columns: list[ColumnDescription] = field(default_factory=list)
    sample_rows: list[list[Any]] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)   # observations like "double header"


@dataclass
class SheetSummary:
    """Full structural summary of an Excel workbook."""

    file_name: str
    total_sheets: int
    sheets: list[SheetDescription] = field(default_factory=list)

    def to_dict(self) -> dict:
        """Convert to a JSON-serialisable dictionary."""
        return {
            "file_name": self.file_name,
            "total_sheets": self.total_sheets,
            "sheets": [
                {
                    "name": s.name,
                    "row_count": s.row_count,
                    "col_count": s.col_count,
                    "header_row_index": s.header_row_index,
                    "has_formulas": s.has_formulas,
                    "notes": s.notes,
                    "columns": [
                        {
                            "index": c.index,
                            "name": c.name,
                            "detected_type": c.detected_type,
                            "non_empty_count": c.non_empty_count,
                            "sample_values": [
                                str(v) if v is not None else None
                                for v in c.sample_values
                            ],
                            **({"formula_example": c.formula_example}
                               if c.formula_example else {}),
                        }
                        for c in s.columns
                    ],
                    "sample_rows": [
                        [str(v) if v is not None else None for v in row]
                        for row in s.sample_rows
                    ],
                }
                for s in self.sheets
            ],
        }


class SheetScanner:
    """Scans Excel workbooks to produce compact structural summaries.

    Does NOT load the full dataset — only reads headers, a sample of rows,
    and column type statistics. Safe to run on large files (1000+ row sheets).
    """

    # Month name patterns used to detect wide-monthly column layouts
    _MONTH_NAMES = {
        "january", "february", "march", "april", "may", "june",
        "july", "august", "september", "october", "november", "december",
        "jan", "feb", "mar", "apr", "jun", "jul", "aug", "sep", "oct", "nov", "dec",
    }
    _QUARTER_PATTERN = re.compile(r"^q[1-4]$", re.IGNORECASE)
    _MONTH_NUMBER_PATTERN = re.compile(
        r"^(\d{1,2})\s*[-–]\s*([a-zA-Z]+)$"  # matches "01 - January"
    )

    def scan(self, file_content: bytes, file_name: str) -> SheetSummary:
        """Scan an Excel file and return a compact structural summary.

        Args:
            file_content: Raw Excel bytes
            file_name: Original filename (used for metadata only)

        Returns:
            SheetSummary describing each sheet's structure

        Raises:
            ValueError: If the file cannot be opened as a valid Excel workbook
        """
        try:
            # Load with data_only=True to read formula results (not expressions)
            wb_values = load_workbook(BytesIO(file_content), data_only=True)
            # Also load raw (with formulas) to detect formula cells
            try:
                wb_formulas = load_workbook(BytesIO(file_content), data_only=False)
            except Exception:
                wb_formulas = None

            summary = SheetSummary(
                file_name=file_name,
                total_sheets=len(wb_values.worksheets),
            )

            for sheet_idx, sheet in enumerate(wb_values.worksheets):
                formula_sheet = (
                    wb_formulas.worksheets[sheet_idx]
                    if wb_formulas and sheet_idx < len(wb_formulas.worksheets)
                    else None
                )
                sheet_desc = self._scan_sheet(sheet, formula_sheet)
                summary.sheets.append(sheet_desc)

            wb_values.close()
            if wb_formulas:
                wb_formulas.close()

            logger.info(
                f"SheetScanner: scanned '{file_name}' — "
                f"{summary.total_sheets} sheets"
            )
            return summary

        except Exception as e:
            logger.error(f"SheetScanner error on '{file_name}': {e}")
            raise ValueError(f"Could not scan Excel file '{file_name}': {e}") from e

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    def _scan_sheet(self, sheet, formula_sheet) -> SheetDescription:
        """Scan a single worksheet."""
        row_count = sheet.max_row or 0
        col_count = min(sheet.max_column or 0, MAX_COLS_DESCRIBED)

        if row_count == 0 or col_count == 0:
            return SheetDescription(
                name=sheet.title,
                row_count=0,
                col_count=0,
                header_row_index=1,
                has_formulas=False,
                notes=["Sheet is empty"],
            )

        # Detect special multi-row-header / transposed layouts BEFORE header detection
        special_notes = self._detect_special_layouts(sheet, row_count, col_count)

        # Detect header row (first non-empty row within first 5)
        header_row_idx, headers = self._detect_header_row(sheet, col_count, row_count)
        data_start = header_row_idx + 1

        notes: list[str] = list(special_notes)

        # Detect formula cells and collect examples
        formula_examples: dict[int, str] = {}  # col_index -> first formula string
        has_formulas = False
        if formula_sheet:
            has_formulas, formula_examples = self._detect_formulas(
                formula_sheet, col_count, row_count
            )

        # Infer column types from data rows
        col_type_info = self._infer_column_types(
            sheet, col_count, data_start, row_count, formula_examples
        )

        # Detect if this looks like a wide-monthly layout
        monthly_cols = [
            h for h in headers
            if self._is_month_column(h)
        ]
        if len(monthly_cols) >= 3:
            notes.append(
                f"wide_monthly layout detected: "
                f"{len(monthly_cols)} month columns "
                f"({monthly_cols[0]} … {monthly_cols[-1]})"
            )

        # Detect quarterly columns
        quarterly_cols = [h for h in headers if self._QUARTER_PATTERN.match(str(h))]
        if quarterly_cols:
            notes.append(f"quarterly columns detected: {quarterly_cols}")

        # Check for double header (row 1 values look like formulas referencing other cells)
        if header_row_idx == 2:
            notes.append("double header detected: row 1 may contain formula references, row 2 used as header")

        # Sample rows (values only, after header)
        sample_rows = self._extract_sample_rows(sheet, data_start, col_count)

        # Build column descriptions
        columns = []
        for ci, header in enumerate(headers[:col_count]):
            type_info = col_type_info[ci] if ci < len(col_type_info) else {}
            col_desc = ColumnDescription(
                index=ci,
                name=str(header) if header is not None else f"Column_{ci + 1}",
                detected_type=type_info.get("type", "unknown"),
                sample_values=type_info.get("samples", []),
                non_empty_count=type_info.get("non_empty", 0),
                formula_example=formula_examples.get(ci),
            )
            columns.append(col_desc)

        return SheetDescription(
            name=sheet.title,
            row_count=row_count,
            col_count=col_count,
            header_row_index=header_row_idx,
            has_formulas=has_formulas,
            columns=columns,
            sample_rows=sample_rows,
            notes=notes,
        )

    def _detect_header_row(
        self, sheet, col_count: int, row_count: int
    ) -> tuple[int, list]:
        """Find the first row that contains meaningful text headers.

        Returns:
            (1-based header row index, list of header values)
        """
        for row_idx in range(1, min(6, row_count + 1)):
            row_vals = []
            for col_idx in range(1, col_count + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                row_vals.append(cell.value)

            # A valid header row has at least 2 non-empty string-like values
            non_empty = [v for v in row_vals if v is not None and str(v).strip()]
            text_like = [
                v for v in non_empty
                if isinstance(v, str) and not str(v).startswith("=")
            ]
            if len(text_like) >= 2:
                return row_idx, row_vals

        # Fallback: use row 1 regardless
        row_vals = [
            sheet.cell(row=1, column=ci).value
            for ci in range(1, col_count + 1)
        ]
        return 1, row_vals

    def _detect_formulas(
        self, formula_sheet, col_count: int, row_count: int
    ) -> tuple[bool, dict[int, str]]:
        """Detect formula cells and collect one example per column."""
        has_formulas = False
        formula_examples: dict[int, str] = {}
        scan_limit = min(row_count, 50)

        for row_idx in range(1, scan_limit + 1):
            for col_idx in range(1, col_count + 1):
                cell = formula_sheet.cell(row=row_idx, column=col_idx)
                if (
                    isinstance(cell.value, str)
                    and cell.value.startswith("=")
                    and (col_idx - 1) not in formula_examples
                    and len(formula_examples) < MAX_FORMULA_EXAMPLES * col_count
                ):
                    has_formulas = True
                    formula_examples[col_idx - 1] = cell.value[:120]  # truncate long formulas

        return has_formulas, formula_examples

    def _infer_column_types(
        self,
        sheet,
        col_count: int,
        data_start: int,
        row_count: int,
        formula_examples: dict[int, str],
    ) -> list[dict]:
        """Infer column types by sampling data rows."""
        col_values: list[list] = [[] for _ in range(col_count)]
        scan_end = min(data_start + TYPE_SCAN_ROWS, row_count + 1)

        for row_idx in range(data_start, scan_end):
            for ci in range(col_count):
                cell = sheet.cell(row=row_idx, column=ci + 1)
                if cell.value is not None:
                    col_values[ci].append(cell.value)

        results = []
        for ci in range(col_count):
            vals = col_values[ci]
            non_empty = len(vals)

            if ci in formula_examples:
                detected_type = "formula"
            elif not vals:
                detected_type = "empty"
            else:
                num_count = sum(1 for v in vals if isinstance(v, (int, float)))
                str_count = sum(1 for v in vals if isinstance(v, str))
                # Check for dates (openpyxl returns datetime objects)
                import datetime
                date_count = sum(1 for v in vals if isinstance(v, (datetime.datetime, datetime.date)))

                if date_count > len(vals) * 0.5:
                    detected_type = "date"
                elif num_count > len(vals) * 0.7:
                    detected_type = "numeric"
                elif str_count > len(vals) * 0.7:
                    detected_type = "text"
                else:
                    detected_type = "mixed"

            # Collect a few sample values (skip None, limit to 3)
            samples = [
                v for v in vals[:SAMPLE_ROW_COUNT * 2]
                if v is not None
            ][:3]

            results.append({
                "type": detected_type,
                "non_empty": non_empty,
                "samples": samples,
            })

        return results

    def _extract_sample_rows(
        self, sheet, data_start: int, col_count: int
    ) -> list[list[Any]]:
        """Extract the first SAMPLE_ROW_COUNT data rows."""
        sample_rows = []
        for row_idx in range(data_start, data_start + SAMPLE_ROW_COUNT):
            row_vals = [
                sheet.cell(row=row_idx, column=ci + 1).value
                for ci in range(col_count)
            ]
            # Only include rows that have at least one non-None value
            if any(v is not None for v in row_vals):
                sample_rows.append(row_vals)
        return sample_rows

    def _detect_special_layouts(
        self, sheet, row_count: int, col_count: int
    ) -> list[str]:
        """Detect dual_metric_wide_monthly and transposed_financial layouts.

        Returns additional notes to append to the sheet description.
        Called before header detection so it can examine raw rows 1–5.
        """
        import datetime

        notes: list[str] = []
        if row_count < 2 or col_count < 3:
            return notes

        # Read first 5 raw rows for pattern analysis
        raw_rows: list[list] = []
        for r in range(1, min(6, row_count + 1)):
            row_vals = [
                sheet.cell(row=r, column=c).value
                for c in range(1, col_count + 1)
            ]
            raw_rows.append(row_vals)

        # ── Pattern A: dual_metric_wide_monthly ──────────────────────────────
        # Row 0 (sheet row 1): sparse text labels spanning column groups
        # Row 1 (sheet row 2): many columns with duplicate month names
        if len(raw_rows) >= 2:
            row0 = raw_rows[0]
            row1 = raw_rows[1]

            # Sparse labels in row 0: few non-None text values across many columns
            row0_labels = [
                (i, v)
                for i, v in enumerate(row0)
                if v is not None and isinstance(v, str) and v.strip()
            ]
            row1_month_vals = [
                v for v in row1
                if v is not None and self._is_month_column(v)
            ]

            if (
                2 <= len(row0_labels) <= max(2, col_count // 4)
                and len(row1_month_vals) >= 6
            ):
                month_strs = [str(v).strip().lower() for v in row1_month_vals]
                if len(month_strs) > len(set(month_strs)):
                    # Duplicate months → dual_metric layout
                    blocks = []
                    for label_idx, (col_idx, label) in enumerate(row0_labels):
                        end_col = (
                            row0_labels[label_idx + 1][0] - 1
                            if label_idx + 1 < len(row0_labels)
                            else col_count - 1
                        )
                        blocks.append(
                            f"{{label:'{label}', start:{col_idx}, end:{end_col}}}"
                        )
                    notes.append(
                        f"dual_metric_header: metric_blocks=[{', '.join(blocks)}]"
                    )

        # ── Pattern B: transposed_financial ──────────────────────────────────
        # Col 0 or col 1 has >60% non-empty text labels
        # AND first few rows contain year-like or date-like column headers
        for col_idx in range(min(2, col_count)):
            col_vals = [
                sheet.cell(row=r, column=col_idx + 1).value
                for r in range(1, min(TYPE_SCAN_ROWS + 1, row_count + 1))
            ]
            non_empty = [v for v in col_vals if v is not None]
            if not non_empty:
                continue

            text_count = sum(
                1 for v in non_empty
                if isinstance(v, str) and v.strip() and not str(v).startswith("=")
            )
            if text_count / len(non_empty) < 0.6:
                continue

            # Check if other columns in early rows look like dates/years
            date_header_rows: list[int] = []
            for ri, row in enumerate(raw_rows[:4]):
                other_vals = [row[c] for c in range(col_count) if c != col_idx]
                has_date_like = any(
                    isinstance(v, (datetime.datetime, datetime.date))
                    or (isinstance(v, (int, float)) and 2000 <= float(v) <= 2035)
                    or (isinstance(v, str) and re.match(r"^\d{4}", str(v).strip()))
                    for v in other_vals
                    if v is not None
                )
                if has_date_like:
                    date_header_rows.append(ri)

            if date_header_rows:
                col_letter = chr(ord("A") + col_idx)
                notes.append(
                    f"transposed_financial: label_column={col_letter}, "
                    f"date_header_rows={date_header_rows}"
                )
                break  # Only annotate once

        return notes

    def _is_month_column(self, header) -> bool:
        """Return True if a column header looks like a month name or month pattern."""
        if header is None:
            return False
        h = str(header).strip().lower()
        # Exact month names
        if h in self._MONTH_NAMES:
            return True
        # "01 - January" style
        if self._MONTH_NUMBER_PATTERN.match(h):
            return True
        return False
