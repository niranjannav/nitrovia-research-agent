"""Generic data extractor.

Uses a SchemaMapping (produced by SchemaInferenceService) to extract structured
analytics records from any Excel workbook, regardless of format.

Handles time structures:
  - wide_monthly  : one row per record, columns = Jan/Feb/.../Dec
  - wide_weekly   : one row per record, columns = week numbers
  - long_date_col : one row per transaction with an explicit date column
  - quarterly_pivot: columns = Q1/Q2/Q3/Q4
  - annual_only   : single annual total per row

All output is a list of AnalyticsRecord dicts that the analytics_data_service
inserts into the domain-specific SQL table.
"""

import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.services.schema_inference_service import SchemaMapping

logger = logging.getLogger(__name__)

# Month name → number mapping (handles both full names and abbreviations)
_MONTH_MAP: dict[str, int] = {
    "january": 1, "jan": 1,
    "february": 2, "feb": 2,
    "march": 3, "mar": 3,
    "april": 4, "apr": 4,
    "may": 5,
    "june": 6, "jun": 6,
    "july": 7, "jul": 7,
    "august": 8, "aug": 8,
    "september": 9, "sep": 9, "sept": 9,
    "october": 10, "oct": 10,
    "november": 11, "nov": 11,
    "december": 12, "dec": 12,
}

_MONTH_NUM_PATTERN = re.compile(r"^(\d{1,2})\s*[-–]\s*([a-zA-Z]+)$")  # "01 - January"
_SHEET_YEAR_PATTERN = re.compile(r"(\d{4})$")                           # trailing 4-digit year


@dataclass
class AnalyticsRecord:
    """Domain-agnostic record produced by the extractor.

    Fields that are not relevant to a given domain will be None.
    The analytics_data_service maps these to domain-specific DB columns.
    """

    record_date: date                   # date this record belongs to
    product_code: str | None = None
    product_name: str | None = None
    customer_code: str | None = None
    customer_name: str | None = None
    channel: str | None = None
    region: str | None = None
    salesperson: str | None = None
    team: str | None = None
    category: str | None = None
    source_name: str | None = None      # for production (milk source)

    # Numeric metrics — domain services pick which are relevant
    quantity_units: float | None = None
    quantity_litres: float | None = None
    revenue: float | None = None
    cogs: float | None = None
    gross_margin: float | None = None
    operating_cost: float | None = None
    net_profit: float | None = None
    orders_fulfilled: float | None = None
    quantity_received_litres: float | None = None
    quantity_processed_litres: float | None = None
    case_count: float | None = None
    # Finance-specific metrics
    gross_profit: float | None = None
    gross_profit_pct: float | None = None
    other_income: float | None = None
    operating_expenses: float | None = None
    operating_profit: float | None = None
    ebit: float | None = None
    ebitda: float | None = None
    net_income: float | None = None
    litres_sold: float | None = None
    revenue_per_litre: float | None = None
    cost_per_litre: float | None = None
    total_assets: float | None = None
    total_liabilities: float | None = None
    total_equity: float | None = None
    operating_cash_flow: float | None = None
    investing_cash_flow: float | None = None
    financing_cash_flow: float | None = None
    net_cash_flow: float | None = None

    extra: dict[str, Any] = field(default_factory=dict)  # unmapped columns


class GenericDataExtractor:
    """Extracts AnalyticsRecord objects from an Excel file using a SchemaMapping.

    Usage:
        extractor = GenericDataExtractor()
        records = extractor.extract(file_bytes, mapping)
    """

    def extract(
        self,
        file_content: bytes,
        mapping: SchemaMapping,
    ) -> list[AnalyticsRecord]:
        """Extract records from the Excel file using the saved SchemaMapping.

        Args:
            file_content: Raw Excel bytes
            mapping: SchemaMapping produced by SchemaInferenceService

        Returns:
            List of AnalyticsRecord objects (one per data cell/row depending on time structure)
        """
        wb = load_workbook(BytesIO(file_content), data_only=True)
        records: list[AnalyticsRecord] = []

        for sheet_name in mapping.source_sheets:
            if sheet_name not in wb.sheetnames:
                # Try to find sheets matching a year pattern
                matched = [
                    s for s in wb.sheetnames
                    if s.replace(sheet_name.rstrip("0123456789"), "").isdigit()
                    or sheet_name in s
                ]
                if matched:
                    for s in matched:
                        sheet = wb[s]
                        sheet_records = self._extract_from_sheet(sheet, mapping)
                        records.extend(sheet_records)
                else:
                    logger.warning(
                        f"Source sheet '{sheet_name}' not found in workbook. "
                        f"Available: {wb.sheetnames}"
                    )
                continue

            sheet = wb[sheet_name]
            sheet_records = self._extract_from_sheet(sheet, mapping)
            records.extend(sheet_records)

        wb.close()

        logger.info(
            f"GenericDataExtractor: extracted {len(records)} records "
            f"from {len(mapping.source_sheets)} sheet(s)"
        )
        return records

    # ── Sheet-level dispatch ─────────────────────────────────────────────────

    def _extract_from_sheet(
        self, sheet, mapping: SchemaMapping
    ) -> list[AnalyticsRecord]:
        time_type = mapping.time_structure.type
        if time_type == "wide_monthly":
            return self._extract_wide(sheet, mapping, period="monthly")
        elif time_type == "wide_weekly":
            return self._extract_wide(sheet, mapping, period="weekly")
        elif time_type == "quarterly_pivot":
            return self._extract_wide(sheet, mapping, period="quarterly")
        elif time_type == "long_date_col":
            return self._extract_long_date(sheet, mapping)
        elif time_type == "annual_only":
            return self._extract_annual(sheet, mapping)
        elif time_type == "dual_metric_wide_monthly":
            return self._extract_dual_metric(sheet, mapping)
        elif time_type == "transposed_financial":
            return self._extract_transposed_financial(sheet, mapping)
        else:
            logger.warning(f"Unknown time_structure.type '{time_type}', skipping sheet")
            return []

    # ── Wide format (wide_monthly / wide_weekly / quarterly_pivot) ───────────

    def _extract_wide(
        self,
        sheet,
        mapping: SchemaMapping,
        period: str,
    ) -> list[AnalyticsRecord]:
        """Extract from wide-format sheets where each time period is a column."""
        headers, col_index = self._read_headers(sheet, mapping.header_row)
        if not headers:
            return []

        # Determine year from sheet name
        year = self._year_from_sheet_name(sheet.title, mapping)

        # Map time-period columns to dates
        time_col_dates = self._map_time_columns(
            mapping.time_structure.columns, year, period, col_index
        )

        if not time_col_dates:
            logger.warning(
                f"No time columns matched in sheet '{sheet.title}'. "
                f"Expected columns: {mapping.time_structure.columns[:5]}"
            )
            return []

        records: list[AnalyticsRecord] = []

        for row_idx in range(mapping.header_row + 1, (sheet.max_row or 0) + 1):
            row = self._read_row(sheet, row_idx, len(headers))
            if not any(v is not None for v in row):
                continue  # skip empty rows

            # Build the dimensional fields (product, customer, salesperson, etc.)
            dims = self._extract_dims(row, col_index, mapping)

            # Apply derived fields (e.g. channel from customer_code prefix)
            self._apply_derived_fields(dims, mapping)

            # One record per time column
            for col_header, rec_date in time_col_dates.items():
                col_idx = col_index.get(col_header)
                if col_idx is None:
                    continue
                value = self._safe_numeric(row[col_idx] if col_idx < len(row) else None)
                if value is None or value == 0:
                    continue  # skip zero/blank cells to keep table lean

                rec = AnalyticsRecord(record_date=rec_date, **dims)
                # Assign value to the primary metric
                self._assign_primary_metric(rec, value, mapping)
                records.append(rec)

        return records

    # ── Long date format (one row per transaction) ───────────────────────────

    def _extract_long_date(
        self,
        sheet,
        mapping: SchemaMapping,
    ) -> list[AnalyticsRecord]:
        """Extract from sheets with an explicit date column."""
        headers, col_index = self._read_headers(sheet, mapping.header_row)
        if not headers:
            return []

        date_col_name = (
            mapping.time_structure.columns[0]
            if mapping.time_structure.columns
            else None
        )
        if not date_col_name or date_col_name not in col_index:
            logger.warning(f"Date column '{date_col_name}' not found in sheet '{sheet.title}'")
            return []

        records: list[AnalyticsRecord] = []
        for row_idx in range(mapping.header_row + 1, (sheet.max_row or 0) + 1):
            row = self._read_row(sheet, row_idx, len(headers))
            if not any(v is not None for v in row):
                continue

            raw_date = row[col_index[date_col_name]] if col_index[date_col_name] < len(row) else None
            rec_date = self._parse_date(raw_date)
            if rec_date is None:
                continue

            dims = self._extract_dims(row, col_index, mapping)
            self._apply_derived_fields(dims, mapping)

            rec = AnalyticsRecord(record_date=rec_date, **dims)

            # For long-date format, assign all numeric role columns
            for std_field, col_header in mapping.column_roles.items():
                if col_header not in col_index:
                    continue
                value = self._safe_numeric(row[col_index[col_header]])
                if value is not None:
                    self._set_field(rec, std_field, value)

            records.append(rec)

        return records

    # ── Annual only ──────────────────────────────────────────────────────────

    def _extract_annual(
        self,
        sheet,
        mapping: SchemaMapping,
    ) -> list[AnalyticsRecord]:
        """Extract from sheets with a single annual total per row."""
        year = self._year_from_sheet_name(sheet.title, mapping)
        rec_date = date(year, 1, 1)

        headers, col_index = self._read_headers(sheet, mapping.header_row)
        records: list[AnalyticsRecord] = []

        for row_idx in range(mapping.header_row + 1, (sheet.max_row or 0) + 1):
            row = self._read_row(sheet, row_idx, len(headers))
            if not any(v is not None for v in row):
                continue

            dims = self._extract_dims(row, col_index, mapping)
            self._apply_derived_fields(dims, mapping)

            rec = AnalyticsRecord(record_date=rec_date, **dims)

            for std_field, col_header in mapping.column_roles.items():
                if col_header not in col_index:
                    continue
                value = self._safe_numeric(row[col_index[col_header]])
                if value is not None:
                    self._set_field(rec, std_field, value)

            records.append(rec)

        return records

    # ── Dual-metric wide monthly ─────────────────────────────────────────────

    def _extract_dual_metric(
        self,
        sheet,
        mapping: SchemaMapping,
    ) -> list[AnalyticsRecord]:
        """Extract from sheets with a multi-row header and duplicate month columns.

        Layout:
          Row 1: sparse metric labels (e.g. "Qty" spanning cols 6-17, "Lts" cols 19-30)
          Row 2: actual column names with duplicate month names
          Row 3+: data rows

        The mapping's time_structure.metric_blocks tells us which column ranges
        carry which metric. The mapping's header_row (1-based) points to row 2.
        """
        ts = mapping.time_structure
        metric_blocks = ts.metric_blocks or []

        if not metric_blocks:
            logger.warning(
                f"dual_metric_wide_monthly: no metric_blocks in time_structure "
                f"for sheet '{sheet.title}', falling back to wide_monthly"
            )
            return self._extract_wide(sheet, mapping, period="monthly")

        year = self._year_from_sheet_name(sheet.title, mapping)
        headers, col_index = self._read_headers(sheet, mapping.header_row)

        records: list[AnalyticsRecord] = []
        # Key: (product_code, customer_code, record_date) → AnalyticsRecord
        merged: dict[tuple, AnalyticsRecord] = {}

        for row_idx in range(mapping.header_row + 1, (sheet.max_row or 0) + 1):
            row = self._read_row(sheet, row_idx, len(headers))
            if not any(v is not None for v in row):
                continue

            dims = self._extract_dims(row, col_index, mapping)
            self._apply_derived_fields(dims, mapping)

            for block in metric_blocks:
                maps_to: str = block.get("maps_to", "quantity_units")
                start_col: int = block.get("start_col", block.get("start", 0))
                end_col: int = block.get("end_col", block.get("end", 0))

                # Build month→col_index map from the block's column range
                # The header row has duplicate month names; use positional index
                block_month_dates: list[tuple[int, date]] = []
                month_counter: dict[int, int] = {}  # month_num → occurrence count

                for ci in range(start_col, min(end_col + 1, len(headers))):
                    h = headers[ci]
                    if h is None:
                        continue
                    d = self._parse_period_header(str(h), year, "monthly")
                    if d:
                        # Handle duplicate month names by counting occurrences
                        key = d.month
                        month_counter[key] = month_counter.get(key, 0) + 1
                        block_month_dates.append((ci, d))

                for ci, rec_date in block_month_dates:
                    value = self._safe_numeric(row[ci] if ci < len(row) else None)
                    if value is None or value == 0:
                        continue

                    merge_key = (
                        dims.get("product_code"),
                        dims.get("customer_code"),
                        rec_date,
                    )

                    if merge_key not in merged:
                        merged[merge_key] = AnalyticsRecord(
                            record_date=rec_date, **dims
                        )

                    rec = merged[merge_key]
                    self._set_field(rec, maps_to, value)

        return list(merged.values())

    # ── Transposed financial ─────────────────────────────────────────────────

    def _extract_transposed_financial(
        self,
        sheet,
        mapping: SchemaMapping,
    ) -> list[AnalyticsRecord]:
        """Extract from transposed financial sheets (rows = line items, cols = dates).

        The mapping's time_structure.date_header_row (1-based) points to the row
        that contains date/year values as column headers.
        The mapping's time_structure.label_column ('B') identifies the column with
        line item labels.
        The mapping's time_structure.row_label_map maps label strings to standard
        field names (e.g. {"Total Sales Revenue": "revenue"}).
        """
        ts = mapping.time_structure
        label_col_letter = ts.label_column or "B"
        row_label_map: dict[str, str] = ts.row_label_map or {}
        date_header_row = ts.date_header_row or mapping.header_row

        # Convert label column letter to 0-based index
        label_col_idx = ord(label_col_letter.upper()) - ord("A")

        max_col = sheet.max_column or 0
        max_row = sheet.max_row or 0

        if max_col == 0 or max_row == 0:
            return []

        # Build col_index → date map from the date header row
        col_date_map: dict[int, date] = {}
        for ci in range(0, max_col):
            cell_val = sheet.cell(row=date_header_row, column=ci + 1).value
            if cell_val is None:
                continue
            d = self._parse_date(cell_val)
            if d:
                col_date_map[ci] = d
            elif isinstance(cell_val, (int, float)) and 2000 <= float(cell_val) <= 2035:
                # Year as integer — use Jan 1 of that year
                col_date_map[ci] = date(int(cell_val), 1, 1)

        if not col_date_map:
            logger.warning(
                f"transposed_financial: no date columns found in row {date_header_row} "
                f"of sheet '{sheet.title}'"
            )
            return []

        # Build label→row_index map by scanning the label column
        label_row_map: dict[str, int] = {}
        for ri in range(1, max_row + 1):
            label_val = sheet.cell(row=ri, column=label_col_idx + 1).value
            if label_val is not None and isinstance(label_val, str) and label_val.strip():
                label_row_map[label_val.strip()] = ri

        # For each date column, build one record from all mapped row labels
        date_records: dict[date, dict] = {}

        for std_field, label in row_label_map.items():
            # Find the row with this label
            row_idx = None
            for raw_label, ri in label_row_map.items():
                if raw_label.strip().lower() == label.strip().lower():
                    row_idx = ri
                    break
                # Fuzzy: label is a substring
                if label.strip().lower() in raw_label.lower():
                    row_idx = ri
                    break
            if row_idx is None:
                continue

            for ci, rec_date in col_date_map.items():
                value = self._safe_numeric(
                    sheet.cell(row=row_idx, column=ci + 1).value
                )
                if value is None:
                    continue

                if rec_date not in date_records:
                    date_records[rec_date] = {"record_date": rec_date}
                date_records[rec_date][std_field] = value

        # Convert to AnalyticsRecord objects
        records: list[AnalyticsRecord] = []
        for rec_dict in date_records.values():
            rec_date = rec_dict.pop("record_date")
            rec = AnalyticsRecord(record_date=rec_date)
            for field_name, value in rec_dict.items():
                self._set_field(rec, field_name, value)
            records.append(rec)

        logger.info(
            f"transposed_financial: extracted {len(records)} period records "
            f"from sheet '{sheet.title}'"
        )
        return records

    # ── Helpers ──────────────────────────────────────────────────────────────

    def _read_headers(
        self, sheet, header_row: int
    ) -> tuple[list, dict[str, int]]:
        """Read the header row and build a name→column-index lookup."""
        row_data = [
            sheet.cell(row=header_row, column=ci + 1).value
            for ci in range(sheet.max_column or 0)
        ]
        col_index: dict[str, int] = {}
        for ci, val in enumerate(row_data):
            if val is not None:
                key = str(val).strip()
                if key and key not in col_index:
                    col_index[key] = ci
        return row_data, col_index

    def _read_row(self, sheet, row_idx: int, ncols: int) -> list:
        return [
            sheet.cell(row=row_idx, column=ci + 1).value
            for ci in range(ncols)
        ]

    def _year_from_sheet_name(self, sheet_name: str, mapping: SchemaMapping) -> int:
        """Extract year from sheet name suffix or fall back to mapping's year_value."""
        if mapping.time_structure.year_source == "sheet_name_suffix":
            m = _SHEET_YEAR_PATTERN.search(sheet_name)
            if m:
                return int(m.group(1))
        if mapping.time_structure.year_value:
            return mapping.time_structure.year_value
        return datetime.now().year

    def _map_time_columns(
        self,
        time_columns: list[str],
        year: int,
        period: str,
        col_index: dict[str, int],
    ) -> dict[str, date]:
        """Map time column headers to date objects.

        Returns dict of {column_header: date} for columns found in col_index.
        """
        result: dict[str, date] = {}
        for header in time_columns:
            if header not in col_index:
                continue
            d = self._parse_period_header(header, year, period)
            if d:
                result[header] = d
        return result

    def _parse_period_header(
        self, header: str, year: int, period: str
    ) -> date | None:
        """Convert a column header like '01 - January' or 'Q1' to a date."""
        h = header.strip()

        if period in ("monthly", "wide_monthly"):
            # "01 - January" style
            m = _MONTH_NUM_PATTERN.match(h)
            if m:
                month_num = int(m.group(1))
                return date(year, month_num, 1)
            # Plain month name
            lower = h.lower()
            if lower in _MONTH_MAP:
                return date(year, _MONTH_MAP[lower], 1)

        elif period in ("quarterly", "quarterly_pivot"):
            q_map = {"q1": 1, "q2": 4, "q3": 7, "q4": 10}
            lower = h.lower()
            if lower in q_map:
                return date(year, q_map[lower], 1)

        elif period in ("weekly", "wide_weekly"):
            # Week columns like "W1", "Wk01", or "Week 1"
            w_match = re.match(r"(?:week|wk|w)[\s.]?(\d{1,2})", h, re.IGNORECASE)
            if w_match:
                week_num = int(w_match.group(1))
                try:
                    return date.fromisocalendar(year, week_num, 1)
                except ValueError:
                    return None

        return None

    def _extract_dims(
        self, row: list, col_index: dict[str, int], mapping: SchemaMapping
    ) -> dict:
        """Extract all non-metric dimensional fields from a row."""
        NON_DIM_FIELDS = {
            "quantity_units", "quantity_litres", "revenue", "cogs",
            "gross_margin", "operating_cost", "net_profit",
            "orders_fulfilled", "quantity_received_litres",
            "quantity_processed_litres", "case_count",
        }
        dims: dict = {}
        for std_field, col_header in mapping.column_roles.items():
            if std_field in NON_DIM_FIELDS:
                continue
            if col_header not in col_index:
                continue
            idx = col_index[col_header]
            val = row[idx] if idx < len(row) else None
            dims[std_field] = str(val).strip() if val is not None else None
        return dims

    def _apply_derived_fields(
        self, dims: dict, mapping: SchemaMapping
    ) -> None:
        """Mutate dims in-place to apply derived field logic."""
        for field_name, df in mapping.derived_fields.items():
            source_value = dims.get(df.from_col)
            if source_value is None:
                # Try by column role
                source_col = mapping.column_roles.get(df.from_col)
                source_value = dims.get(source_col) if source_col else None
            if source_value is None:
                continue

            source_str = str(source_value)

            if df.method in ("prefix_2chars", "prefix_Nchars"):
                n = df.prefix_length if df.method == "prefix_Nchars" and df.prefix_length else 2
                prefix = source_str[:n].upper()
                if df.map and prefix in df.map:
                    dims[field_name] = df.map[prefix]
                else:
                    dims[field_name] = prefix

            elif df.method == "regex" and df.pattern:
                m = re.search(df.pattern, source_str)
                if m:
                    captured = m.group(1) if m.lastindex else m.group(0)
                    if df.map and captured in df.map:
                        dims[field_name] = df.map[captured]
                    else:
                        dims[field_name] = captured

            elif df.method == "constant" and df.value is not None:
                dims[field_name] = df.value

    def _assign_primary_metric(
        self, rec: AnalyticsRecord, value: float, mapping: SchemaMapping
    ) -> None:
        """Set the primary numeric metric on a record based on domain."""
        domain = mapping.domain
        if domain == "sales":
            # Check if this sheet is a revenue sheet based on its metric roles
            if "revenue" in mapping.column_roles:
                rec.revenue = value
            else:
                rec.quantity_units = value
        elif domain == "production":
            rec.quantity_processed_litres = value
        elif domain == "qa":
            rec.case_count = value
        elif domain == "finance":
            rec.revenue = value
        else:
            rec.quantity_units = value

    def _set_field(self, rec: AnalyticsRecord, field_name: str, value: Any) -> None:
        """Set a field on a record if it exists."""
        if hasattr(rec, field_name):
            setattr(rec, field_name, value)
        else:
            rec.extra[field_name] = value

    def _safe_numeric(self, value: Any) -> float | None:
        """Convert a value to float, returning None for non-numeric/zero."""
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            cleaned = value.replace(",", "").replace(" ", "").strip()
            try:
                return float(cleaned)
            except ValueError:
                return None
        return None

    def _parse_date(self, value: Any) -> date | None:
        """Parse various date formats to a date object."""
        if value is None:
            return None
        if isinstance(value, date) and not isinstance(value, datetime):
            return value
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, str):
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
                try:
                    return datetime.strptime(value.strip(), fmt).date()
                except ValueError:
                    continue
        return None
